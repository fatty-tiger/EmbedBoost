import json


def reader(input_file, encoding='utf8', fmt='raw'):
    if fmt not in ["raw", "jsonl"]:
        raise ValueError("unsupported format")
    with open(input_file, encoding=encoding) as f:
        if fmt == 'jsonl' or input_file.endswith(".jsonl"):
            for idx, line in enumerate(f):
                d = json.loads(line.strip())
                yield idx, d
        else:
            for idx, line in enumerate(f):
                yield idx, line.strip()


def excel_reader(fpath, sheet_names):
    from openpyxl import load_workbook
    wb = load_workbook(filename=fpath)
    for sheet_name in sheet_names:
        if sheet_name not in wb:
            continue
        ws = wb[sheet_name]
        row_num, col_num = ws.max_row, ws.max_column
        for row in range(1, row_num + 1):
            item = {}
            item['sheet'] = sheet_name
            for col in range(1, col_num + 1):
                key = ws.cell(1, col).value
                item[key] = ws.cell(row, col) .value
            yield item


def writer(datas, output_fpath, fmt='raw', encoding='utf8'):
    if not isinstance(datas, list):
        raise TypeError("data must be a list")
    if fmt not in ['raw', 'jsonl', 'tsv']:
        raise TypeError("fmt invalid")
    
    def get_check_func(fmt):
        if fmt == 'jsonl':
            return lambda x: isinstance(x, dict)
        else:
            return lambda x: isinstance(x, str)

    def get_write_func(fmt):
        if fmt == 'jsonl':
            return lambda x: json.dumps(x, ensure_ascii=False)
        else:
            return lambda x: x

    check_func = get_check_func(fmt)
    write_func = get_write_func(fmt)
    with open(output_fpath, 'w', encoding=encoding) as wr:
        for item in datas:
            if not check_func(item):
                raise ValueError("invalid item")
            wr.write(write_func(item) + '\n')


def batch_generator(input_fpath_or_list, batch_size):
    idx, batch = 0, []
    if isinstance(input_fpath_or_list, str):
        f = open(input_fpath_or_list, encoding='utf8')
        for line in f:
            batch.append(line.strip())
            if len(batch) == batch_size:
                yield idx, batch
                batch = []
                idx += 1
        f.close()
    elif isinstance(input_fpath_or_list, list):
        for x in input_fpath_or_list:
            batch.append(x)
            if len(batch) == batch_size:
                yield idx, batch
                batch = []
                idx += 1

    if len(batch) > 0:
        yield idx, batch


def save_jsonl(datas, output_fpath): 
    if not output_fpath.endswith('.jsonl'):
        raise TypeError("output_fpath is not jsonl format!")
    with open(output_fpath, "w") as wr:
        for d in datas:
            wr.write(json.dumps(d, ensure_ascii=False) + '\n')


def jsonl_to_excel(input_fpath_or_list, output_fpath, mode="w", header_mapping=None, sheet_name=None):
    if not output_fpath.endswith('.xlsx'):
        raise TypeError("output_fpath is not xlsx format!")
 
    datas = None
    if isinstance(input_fpath_or_list, str):
        if not input_fpath_or_list.endswith('.jsonl'):
            raise TypeError("input_fpath is not jsonl format!")
        # 将json数据转为pandas DataFrame  
        datas = []
        for  _, d in reader(input_fpath_or_list, fmt='jsonl'):
            if header_mapping:
                new_d = {header_mapping.get(k, k): v for k, v in d.items()}
                datas.append(new_d)
            else:
                datas.append(d)
    elif isinstance(input_fpath_or_list, list):
        datas = []
        for d in input_fpath_or_list:
            if header_mapping:
                new_d = {header_mapping.get(k, k): v for k, v in d.items()}
                datas.append(new_d)
            else:
                datas.append(d)
    else:
        raise TypeError("input_fpath invalid format!")

    import pandas as pd
    df = pd.DataFrame(datas)
    with pd.ExcelWriter(output_fpath, mode=mode) as writer:
        if sheet_name is None:
            df.to_excel(writer, index=False)
        else:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
